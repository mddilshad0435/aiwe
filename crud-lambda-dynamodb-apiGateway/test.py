from openpyxl import load_workbook, Workbook
import time
import requests
from datetime import datetime
from requests_html import AsyncHTMLSession, HTMLSession
import time
import pytz
import os
import sys
import json
import logging
from googleapiclient.http import MediaFileUpload
import re
import pickle
import random
import os.path
import os
import uuid
from Google import Create_Service
import asyncio, aiohttp


class Bot:

    def __init__(self):
        self.email = 'mailto:yuluxuryholdings@gmail.com'
        self.password = 'e9ArKFQT'
        self.session = HTMLSession()
    
    def drive_config(self):
        self.CLIENT_SECRET_FILE = 'client_secret_file.json'
        self.API_NAME = 'drive'
        self.API_VERSION = 'v3'
        self.SCOPES = ['https://www.googleapis.com/auth/drive',
                    'https://www.googleapis.com/auth/drive.file',
                    'https://www.googleapis.com/auth/spreadsheets',
                    'https://www.googleapis.com/auth/drive.appdata',
                    'https://www.googleapis.com/auth/drive.activity', # Allows read and write access to the Drive Activity API.
                    'https://www.googleapis.com/auth/drive.scripts', # access to apps script
                    ]
        self.service = Create_Service(self.CLIENT_SECRET_FILE, self.API_NAME, self.API_VERSION, self.SCOPES)

    def get_login_page(self):
        '''Tries to get login page, if error, logs to file.'''
        logging.basicConfig(filename='get_login_errors.log', filemode='a', \
            format='%(asctime)s - %(message)s', datefmt='%d-%b-%y %H:%M:%S')
        try:
            self.login_page = self.session.get('https://www.starbuyers-global-auction.com/login')
        except Exception as e:
            logging.exception(f'Exception occurred: {e}', exc_info=True)
            time.sleep(random.randint(3, 6))
            self.get_login_page()



    def get_login_token(self):
        self.login_page_soup = self.login_page.html
        self.hidden_input = self.login_page_soup.find('input[name=_token]')
        self.token_element = self.hidden_input[0]
        self.hidden_dict = self.token_element.attrs
        self.token = self.hidden_dict['value']
        print(f'Login token: {self.token}')
        self.login_data = {'_token': self.token, 'email': self.email, 'password': self.password}


    def post_login_info(self):
        logging.basicConfig(filename='post_login_errors.log', filemode='a', \
            format='%(asctime)s - %(message)s', datefmt='%d-%b-%y %H:%M:%S')
        try:
            self.session.post('https://www.starbuyers-global-auction.com/login', data=self.login_data)
        except Exception as e:
            logging.exception(f'Exception occurred: {e}', exc_info=True)
            time.sleep(random.randint(3, 6))
            self.post_login_info()

    def read_urls(self):
        self.un_urls = []
        self.wb = load_workbook('Beg_Nov_URL_2.xlsx')
        self.sheet = self.wb['Sheet1']
        self.end_row = self.sheet.max_row
        self.start_row = self.sheet.min_row
        for row in self.sheet[f'A{self.start_row}': f'A{self.end_row}']:
            for cell in row:
                self.un_urls.append(str(cell.value))
        self.urls = [url for url in self.un_urls if url != 'None']

    async def fetch_url(self, url, sheet, count):
            print("in fetch urls", url)
            print(sheet, count)
        
   
            product_page = self.session.get(url)
            product_soup = product_page.html
            
            end_script = product_soup.xpath('/html/body/script[3]')[0].text
            lot_pattern = "'[0-9]+-[0-9]+'"
            lot_result = re.findall(lot_pattern, end_script)
            print("lot_result",lot_result)
            un_lot_no = lot_result[0]
            lot_no = ''
            for letter in un_lot_no:
                if letter != "'":
                    lot_no += letter
            sheet[f'A{count}'] = lot_no
            brand = product_soup.xpath('//*[@id="vue-item-detail"]/div[1]/div[1]/div/div[10]/div/div[1]/dl/dd[4]/text()')[0]
            sheet[f'B{count}'] = brand
            description = product_soup.xpath('//h2[@class="p-title-large u-mb-2"]/text()')[0].strip()
            print(description)
            sheet[f'C{count}'] = description
            condition = product_soup.xpath('//*[@id="vue-item-detail"]/div[1]/div[1]/div/div[2]/p[2]/span/text()')[0]
            sheet[f'D{count}'] = condition
            photos_pattern = 'image_urls: .+]'
            photos_result = re.findall(photos_pattern, end_script)
            photos_string = photos_result[0]
            new_string = ''
            for char in photos_string:
                if char == '\\':
                    continue
                new_string += char
            format_pattern = '".+"'
            to_format_urls = re.findall(format_pattern, new_string)
            string_to_make_list = to_format_urls[0]
            unformatted_list = string_to_make_list.split(',')
            image_urls = []
            for url in unformatted_list:
                new_url = ''
                for char in url:
                    if char == '"':
                        continue
                    new_url += char
                image_urls.append(new_url)
            sheet[f'E{count}'] = image_urls[0]
            file_metadata = {
                'name': lot_no,
                'mimeType': 'application/vnd.google-apps.folder',
                'parents': '1XPjjKU4Dr72BSJMIzfqeZOrpUHxrDAsz'
            }
            
            folder = self.service.files().create(body=file_metadata,
                                                fields='id').execute()
            folder_id = folder.get('id')
            permissions = {
                "role": "writer",
                "type": "anyone"
            }
            self.service.permissions().create(
                fileId=folder_id,
                body=permissions
            ).execute()
            for url in image_urls:
                
                image_name = f'{str(uuid.uuid4())}.jpg'
                img_data = requests.get(url).content
                with open(image_name, 'wb') as file:
                    file.write(img_data)
                
                image_metadata = {
                    'name': image_name,
                    'parents': [folder_id]
                }
                media = MediaFileUpload(image_name,
                                        mimetype='image/jpeg')
                
                file = self.service.files().create(
                    body=image_metadata,
                    media_body=media,
                    fields='id').execute()
                file_id = file.get('id')
                self.service.permissions().create(
                    fileId=folder_id,
                    body=permissions
                ).execute()
                os.remove(image_name)
            sheet[f'F{count}'] = f'https://drive.google.com/drive/u/0/folders/{folder_id}'
            count+=1
            if count==5:
                return
            
    async def get_urls(self):
        print("in get urls")

        wb = Workbook()
        sheet = wb.active
        sheet['A1'] = 'SKU'
        sheet['B1'] = 'BRAND'
        sheet['C1'] = 'DESCRIPTION'
        sheet['D1'] = 'CONDITION'
        sheet['E1'] = 'MAIN IMAGE'
        sheet['F1'] = "GOOGLE DRIVE IMAGES"
        count = 2
        start = time.perf_counter()

        all_responses = await asyncio.gather(*[self.fetch_url(url, sheet, count) for url in self.urls[0:3:1]])
        
        end = time.perf_counter()
        print(f'TIME TAKEN: {end - start}')
        wb.save('products69.xlsx')

a = Bot()
a.drive_config()
a.get_login_page()
a.get_login_token()
a.post_login_info()
a.read_urls()
asyncio.run(a.get_urls())


        
    
